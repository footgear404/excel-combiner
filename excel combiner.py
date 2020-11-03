import os
import re
import openpyxl
from openpyxl import Workbook

directory = 'D:/excel/test/'


def buildSvodFile(files):
    workFile = openpyxl.load_workbook(directory + files[0])
    sheet = workFile.active
    for row in sheet.iter_rows():
        yield [cell.value for cell in row]


def workWithFile():
    files = os.listdir(directory)
    wbResult = Workbook()
    wsResult = wbResult.active
    ls = []
    fileName = ""

    lists = list(buildSvodFile(files))

    row = 1
    for it in lists:
        # print(it)
        for i in range(len(it)):
            wsResult.cell(row=row, column=i + 1).value = it[i]
        row += 1

    for i, name in enumerate(files):
        print("Обработка файла " + name)
        workFile = openpyxl.load_workbook(directory + name)
        sheet = workFile.active
        position = 1
        row = 1
        fileName = name[:8]

        for cell in sheet['N']:
            ls.append([cell.value, position])
            position += 1
            # print(str(cell.value) + " " + str(position))

        for j in ls:
            if j[0] is not None:
                wsResult.cell(row=j[1], column=14).value = j[0]
                row += 1
        print(name + " Обработка закончена \n")
    wsResult.cell(row=1, column=14).value = "Прикрепление"
    wbResult.save(directory + str(fileName) + '_СВОД.xlsx')


def svod():
    files = os.listdir(directory)
    workFile = openpyxl.load_workbook(directory + files[-1])
    sheet = workFile.active
    for row in sheet.iter_rows():
        yield [cell.value for cell in row]


def createMOFile():
    filteredItem = ""
    result = {}

    lists = list(svod())

    for lst in lists:
        if lst[13] not in result:
            result[lst[13]] = []
            # print(lst)
        result[lst[13]].append(lst)
    try:
        os.mkdir(directory + "MO")
    except:
        print("Каталог существует")

    for mo in result:
        if result[mo] != "Прикрепление":
            print("-------------" + mo + "-----------------")
            wbMo = Workbook()
            wsMo = wbMo.active
            row = 1
            col = 1
            for i in result[mo]:
                for j in i:
                    wsMo.cell(row=row, column=col).value = j
                    col += 1
                col = 1
                row += 1
                print(i)
            cleanName = re.sub(r'[^а-яА-Я0-9№ ]+', '', mo)
            wbMo.save(directory + "MO/" + cleanName + '.xlsx')

    # print(result.get('ГБУЗ АО «Бюро судебно-медицинской экспертизы»'))
    # for i in result:
    #     print(result[i])
    # print(result.items())
    # print(result["Наименование учреждения"])


def main():
    workWithFile()
    createMOFile()


if __name__ == '__main__':
    main()
